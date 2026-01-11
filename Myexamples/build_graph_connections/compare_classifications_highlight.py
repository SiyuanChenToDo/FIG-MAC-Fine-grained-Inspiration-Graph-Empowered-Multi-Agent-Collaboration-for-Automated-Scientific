"""
Excel分类结果对比高亮脚本
=================================
功能：
1. 读取Excel中的初次标注列(`human_classification`)与二次标注列(`llm_secondary_classification`)
2. 若两列值不同，则将对应单元格背景标红
3. 支持对单个文件或整个目录批量处理
"""

import argparse
import os
from typing import List

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 默认目录（与标注数据目录保持一致）
DEFAULT_BASE_DIR = "Myexamples/build_graph_connections/annotation_data"
TARGET_SHEET_NAME = "标注数据"
FIRST_COL_NAME = "llm_classification"
SECOND_COL_NAME = "llm_secondary_classification"

HIGHLIGHT_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
RESET_FILL = PatternFill(fill_type=None)


def collect_excel_files(path: str) -> List[str]:
    """根据输入路径收集Excel文件列表"""
    if os.path.isfile(path) and path.lower().endswith(".xlsx"):
        if os.path.basename(path).startswith("~$"):
            return []
        return [path]

    excel_files = []
    for root, _, files in os.walk(path):
        for file in files:
            if file.lower().endswith(".xlsx") and not file.startswith("~$"):
                excel_files.append(os.path.join(root, file))
    return sorted(excel_files)


def find_column_index(header_row) -> dict:
    """返回需要关注的列的列索引映射"""
    mapping = {}
    for cell in header_row:
        if cell.value == FIRST_COL_NAME:
            mapping[FIRST_COL_NAME] = cell.column
        elif cell.value == SECOND_COL_NAME:
            mapping[SECOND_COL_NAME] = cell.column
    return mapping


def highlight_differences(excel_path: str) -> bool:
    """对单个Excel文件执行高亮操作"""
    if not os.path.exists(excel_path):
        print(f"[跳过] 文件不存在: {excel_path}")
        return False

    filename = os.path.basename(excel_path)
    if filename.startswith("~$"):
        print(f"[跳过] 临时或锁定文件: {excel_path}")
        return False

    try:
        wb = load_workbook(excel_path)
    except PermissionError:
        print(f"[跳过] 文件被占用，无法读取: {excel_path}")
        return False
    except OSError as err:
        print(f"[跳过] 无法打开文件 {excel_path}，原因: {err}")
        return False
    if TARGET_SHEET_NAME not in wb.sheetnames:
        print(f"[跳过] 未找到工作表‘{TARGET_SHEET_NAME}’: {excel_path}")
        return False

    ws = wb[TARGET_SHEET_NAME]
    header = next(ws.iter_rows(min_row=1, max_row=1))
    col_map = find_column_index(header)

    missing_cols = [col for col in (FIRST_COL_NAME, SECOND_COL_NAME) if col not in col_map]
    if missing_cols:
        print(f"[跳过] 缺少列 {missing_cols}: {excel_path}")
        return False

    first_col_idx = col_map[FIRST_COL_NAME]
    second_col_idx = col_map[SECOND_COL_NAME]

    diff_count = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2):
        total_rows += 1
        first_cell = row[first_col_idx - 1]
        second_cell = row[second_col_idx - 1]

        first_value = str(first_cell.value).strip() if first_cell.value is not None else ""
        second_value = str(second_cell.value).strip() if second_cell.value is not None else ""

        if not first_value and not second_value:
            first_cell.fill = RESET_FILL
            second_cell.fill = RESET_FILL
            continue

        if first_value != second_value:
            first_cell.fill = HIGHLIGHT_FILL
            second_cell.fill = HIGHLIGHT_FILL
            diff_count += 1
        else:
            first_cell.fill = RESET_FILL
            second_cell.fill = RESET_FILL

    wb.save(excel_path)
    print(f"[完成] {excel_path} -> 总行数 {total_rows}，差异 {diff_count}")
    return True


def main():
    parser = argparse.ArgumentParser(description="高亮初次LLM与二次LLM分类差异")
    parser.add_argument(
        "path",
        nargs="?",
        default=DEFAULT_BASE_DIR,
        help="目标Excel文件或目录（默认：annotation_data目录）"
    )
    args = parser.parse_args()

    excel_files = collect_excel_files(args.path)
    if not excel_files:
        print("未找到任何Excel文件。请确认路径是否正确。")
        return

    print(f"共找到 {len(excel_files)} 个Excel文件，开始对比...")
    for excel_path in excel_files:
        highlight_differences(excel_path)

    print("全部处理完成。")


if __name__ == "__main__":
    main()
